import os

import pandas as pd
from flask import request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation


def format_excel_file(file_path):
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2
    wb.save(file_path)


def adjust_column_width(ws):
    for column in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2


def add_border(ws, cell_range):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border


def create_template_excel(wb, sheet_name, class_code, class_name, common_class_info, is_care=False):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["STT", "Mã sinh viên", "Họ tên sinh viên", "Lớp", "Email", "Lý do nợ môn", "Ghi chú"])

        ws.merge_cells('A1:G1')
        ws['A1'] = "DANH SÁCH SINH VIÊN NỢ MÔN"
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A1'].font = Font(size=16, bold=True)

        ws.merge_cells('A2:G2')
        ws[
            'A2'] = f'(Môn: {common_class_info["subjectCode"]} - {common_class_info["subjectName"]} - {class_code} - {class_name})'
        ws['A2'].alignment = Alignment(horizontal="center")
        ws['A2'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        if is_care:
            headers = ["STT", "Mã sinh viên", "Họ tên sinh viên", "Lớp", "Email", "Số buổi vắng", "Lý do nợ môn",
                       "Ghi chú"]
        else:
            headers = ["STT", "Mã sinh viên", "Họ tên sinh viên", "Lớp", "Email", "Lý do nợ môn", "Ghi chú"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    return wb[sheet_name]


def get_explain_file():
    try:
        data = request.get_json()
        if not data or 'classData' not in data:
            return {'message': 'No data provided'}, 400

        output_dir = os.path.join(os.getcwd(), 'output')
        os.makedirs(output_dir, exist_ok=True)
        local_file_path = os.path.join(output_dir, 'output.xlsx')

        wb = Workbook()
        wb.remove(wb.active)

        for class_info in data['classData']:
            member_data = class_info['memberData']
            score_data = class_info['scoreData']
            common_class_info = class_info['commonClassInfo']
            class_code = class_info['classCode']
            class_name = class_info['className']
            sheet_name = common_class_info['subjectCode']

            ws = create_template_excel(wb, sheet_name, class_code, class_name, common_class_info)

            members_df = pd.DataFrame(member_data)
            scores_df = pd.DataFrame(score_data)

            if members_df.empty or scores_df.empty:
                continue

            members_df_filtered = members_df[members_df['studentId'].isin(scores_df['studentId'])]

            reason_for_failing_options = ["Thi trượt", "Trượt điểm danh", "Đang học", "Khác", "Không có dữ liệu"]
            dv = DataValidation(type="list", formula1=f'"{",".join(reason_for_failing_options)}"', allow_blank=True)
            ws.add_data_validation(dv)

            start_row = ws.max_row + 1
            for row_num, member in enumerate(members_df_filtered.itertuples(), start=start_row - 4):
                student_scores = scores_df[scores_df['studentId'] == member.studentId]
                if not student_scores.empty:
                    status_subject = student_scores['statusSubject'].iloc[0]
                    if status_subject == "0":
                        reason_for_failing = "Thi trượt"
                    elif status_subject == "-1":
                        reason_for_failing = "Trượt điểm danh"
                    elif status_subject == "-2":
                        reason_for_failing = "Đang học"
                    else:
                        reason_for_failing = "Khác"
                else:
                    reason_for_failing = "Không có dữ liệu"

                row_data = [
                    row_num,
                    member.studentCode,
                    member.studentName,
                    f"{member.classCode} - {member.className}",
                    member.studentEmail,
                    reason_for_failing,
                    ""
                ]

                ws.append(row_data)
                dv.add(ws[f'F{ws.max_row}'])

            end_row = ws.max_row
            add_border(ws, f'A{start_row}:G{end_row}')
            add_border(ws, 'A1:G2')
            add_border(ws, 'A4:G4')

            adjust_column_width(ws)
            ws.column_dimensions['A'].width = 5

        wb.save(local_file_path)

        if not os.path.exists(local_file_path):
            return {'message': 'Failed to save the file locally'}, 500

        return send_file(local_file_path, as_attachment=True, download_name='output.xlsx')
    except Exception as e:
        print(e)
        return {'message': str(e)}, 500


def get_care_file():
    try:
        data = request.get_json()
        if not data or 'attendanceData' not in data:
            return {'message': 'No data provided'}, 400

        output_dir = os.path.join(os.getcwd(), 'output')
        os.makedirs(output_dir, exist_ok=True)
        local_file_path = os.path.join(output_dir, 'attendance_output.xlsx')

        wb = Workbook()
        wb.remove(wb.active)

        for class_info in data['attendanceData']:
            attendance_data = class_info['attendanceData']
            common_class_info = class_info['commonClassInfo']
            class_code = class_info['classCode']
            class_name = class_info['className']
            sheet_name = common_class_info['subjectCode']

            ws = create_template_excel(wb, sheet_name, class_code, class_name, common_class_info, is_care=True)

            attendance_df = pd.DataFrame(attendance_data)

            if attendance_df.empty:
                continue

            start_row = ws.max_row + 1
            for row_num, attendance in enumerate(attendance_df.itertuples(), start=start_row - 4):
                reason_for_failing = "Thi trượt" if attendance.status == "0" else "Trượt điểm danh"
                row_data = [
                    row_num,
                    attendance.student_code,
                    attendance.student_name,
                    f"{class_code} - {class_name}",
                    attendance.student_email,
                    f"{attendance.attendance_count}/{attendance.total_session}",
                    reason_for_failing,
                    ""
                ]

                ws.append(row_data)

            end_row = ws.max_row
            add_border(ws, f'A{start_row}:H{end_row}')
            add_border(ws, 'A1:H2')
            add_border(ws, 'A4:H4')

            adjust_column_width(ws)
            ws.column_dimensions['A'].width = 5

        wb.save(local_file_path)

        if not os.path.exists(local_file_path):
            return {'message': 'Failed to save the file locally'}, 500

        return send_file(local_file_path, as_attachment=True, download_name='attendance_output.xlsx')
    except Exception as e:
        print(e)
        return {'message': str(e)}, 500
